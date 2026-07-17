from io import BytesIO

from openpyxl import load_workbook


class ExcelImportService:
    # ==================================================
    # STUDENTS
    # ==================================================

    def read_students(self, file):
        workbook = load_workbook(
            BytesIO(file),
        )

        sheet = workbook.active

        students = []

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            if not any(row):
                continue

            students.append(
                {
                    "first_name": str(row[0]).strip() if row[0] else "",
                    "last_name": str(row[1]).strip() if row[1] else "",
                    "email": str(row[2]).strip().lower() if row[2] else "",
                    "gender": str(row[3]).strip().upper() if row[3] else "",
                    "date_of_birth": row[4],
                    "admission_date": row[5],
                    "class_name": str(row[6]).strip() if row[6] else "",
                }
            )

        return students

    # ==================================================
    # TEACHERS
    # ==================================================

    def read_teachers(self, file):
        workbook = load_workbook(
            BytesIO(file),
        )

        sheet = workbook.active

        teachers = []

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            if not any(row):
                continue

            teachers.append(
                {
                    "first_name": str(row[0]).strip() if row[0] else "",
                    "last_name": str(row[1]).strip() if row[1] else "",
                    "email": str(row[2]).strip().lower() if row[2] else "",
                    "qualification": str(row[3]).strip() if row[3] else "",
                    "specialization": str(row[4]).strip() if row[4] else "",
                    "hire_date": row[5],
                    "class_name": str(row[6]).strip() if row[6] else "",
                }
            )

        return teachers


excel_import_service = ExcelImportService()
