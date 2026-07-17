from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class ExcelTemplateService:
    # =====================================================
    # HELPERS
    # =====================================================

    def _auto_fit(self, worksheet):
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)

            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 4

    def _style_header(self, worksheet):
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        worksheet.freeze_panes = "A2"

    # =====================================================
    # AVAILABLE CLASSES SHEET
    # =====================================================

    def _add_classes_sheet(
        self,
        workbook,
        classes,
    ):
        ws = workbook.create_sheet("Available Classes")

        ws.append(
            [
                "Class Name",
                "Level",
            ]
        )

        self._style_header(ws)

        for school_class in classes:
            ws.append(
                [
                    school_class.name,
                    school_class.level,
                ]
            )

        self._auto_fit(ws)

        # Hide lookup sheet
        ws.sheet_state = "hidden"

    # =====================================================
    # CLASS DROPDOWN
    # =====================================================

    def _add_class_dropdown(
        self,
        worksheet,
        classes,
        column="G",
    ):
        if not classes:
            return

        end_row = len(classes) + 1

        dropdown = DataValidation(
            type="list",
            formula1=f"'Available Classes'!$A$2:$A${end_row}",
            allow_blank=False,
        )

        dropdown.promptTitle = "Class"
        dropdown.prompt = "Select a class."

        dropdown.errorTitle = "Invalid Class"
        dropdown.error = "Please choose a class from the dropdown list."

        worksheet.add_data_validation(dropdown)

        dropdown.add(f"{column}2:{column}5000")

    # =====================================================
    # STUDENT TEMPLATE
    # =====================================================

    def student_template(
        self,
        classes,
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "Students"

        ws.append(
            [
                "first_name",
                "last_name",
                "email",
                "gender",
                "date_of_birth",
                "admission_date",
                "class_name",
            ]
        )

        self._style_header(ws)

        sample_class = classes[0].name if classes else ""

        ws.append(
            [
                "John",
                "Doe",
                "john@example.com",
                "MALE",
                "2012-04-15",
                "2024-09-10",
                sample_class,
            ]
        )

        ws["I1"] = "NOTE"
        ws["J1"] = "Select class_name from the dropdown."

        ws["I1"].font = Font(bold=True)

        self._add_classes_sheet(
            wb,
            classes,
        )

        self._add_class_dropdown(
            ws,
            classes,
            column="G",
        )

        self._auto_fit(ws)

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream

    # =====================================================
    # TEACHER TEMPLATE
    # =====================================================

    def teacher_template(
        self,
        classes,
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "Teachers"

        ws.append(
            [
                "first_name",
                "last_name",
                "email",
                "qualification",
                "specialization",
                "hire_date",
                "class_name",
            ]
        )

        self._style_header(ws)

        sample_class = classes[0].name if classes else ""

        ws.append(
            [
                "Sarah",
                "James",
                "sarah@example.com",
                "B.Ed",
                "Mathematics",
                "2024-01-12",
                sample_class,
            ]
        )

        ws["I1"] = "NOTE"
        ws["J1"] = "Select class_name from the dropdown."

        ws["I1"].font = Font(bold=True)

        self._add_classes_sheet(
            wb,
            classes,
        )

        self._add_class_dropdown(
            ws,
            classes,
            column="G",
        )

        self._auto_fit(ws)

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream

    # =====================================================
    # PARENT TEMPLATE
    # =====================================================

    def parent_template(self):
        wb = Workbook()

        ws = wb.active
        ws.title = "Parents"

        ws.append(
            [
                "first_name",
                "last_name",
                "email",
                "occupation",
                "phone",
            ]
        )

        self._style_header(ws)

        ws.append(
            [
                "David",
                "Johnson",
                "david@example.com",
                "Engineer",
                "+2348012345678",
            ]
        )

        self._auto_fit(ws)

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream


excel_template_service = ExcelTemplateService()
