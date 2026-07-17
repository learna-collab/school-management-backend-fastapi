from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelExportService:
    # =====================================================
    # CREDENTIAL SHEET
    # =====================================================

    def credentials_sheet(
        self,
        users,
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "Credentials"

        headers = [
            "Name",
            "Username",
            "Password",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for user in users:
            ws.append(
                [
                    user["name"],
                    user["username"],
                    user["password"],
                ]
            )

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream

    # =====================================================
    # IMPORT REPORT
    # =====================================================

    def import_report(
        self,
        credentials: list,
        errors: list,
    ):
        wb = Workbook()

        # ==========================================
        # SHEET 1 - SUCCESSFUL IMPORTS
        # ==========================================

        ws = wb.active
        ws.title = "Credentials"

        ws.append(
            [
                "Name",
                "Username",
                "Password",
            ]
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for user in credentials:
            ws.append(
                [
                    user["name"],
                    user["username"],
                    user["password"],
                ]
            )

        # ==========================================
        # SHEET 2 - FAILED IMPORTS
        # ==========================================

        error_sheet = wb.create_sheet(
            "Import Errors",
        )

        error_sheet.append(
            [
                "Excel Row",
                "Name",
                "Email",
                "Reason",
            ]
        )

        for cell in error_sheet[1]:
            cell.font = Font(bold=True)

        for error in errors:
            error_sheet.append(
                [
                    error["row"],
                    error["name"],
                    error["email"],
                    error["reason"],
                ]
            )

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return stream


excel_export_service = ExcelExportService()
