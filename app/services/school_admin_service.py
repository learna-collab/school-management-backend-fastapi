from io import BytesIO

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.repositories.attendance_repository import attendance_repository
from app.repositories.result_repository import result_repository
from app.repositories.school_admin_repository import school_admin_repo
from app.schemas.school_admin import (
    UpdateClassRequest,
    UpdateStudentRequest,
    UpdateSubjectRequest,
    UpdateTeacherRequest,
)


class SchoolAdminService:
    def __init__(self):
        self.repo = school_admin_repo
        self.attendance_repo = attendance_repository
        self.result_repo = result_repository

    # =========================
    # STATS
    # =========================
    async def get_stats(self, db, school_id: str):
        students = await self.repo.count_students(db, school_id)
        teachers = await self.repo.count_teachers(db, school_id)
        classes = await self.repo.count_classes(db, school_id)

        return {
            "students": students,
            "teachers": teachers,
            "classes": classes,
        }

    # =========================
    # STUDENTS
    # =========================
    async def get_students(self, db, school_id: str):
        users = await self.repo.get_students(db, school_id)

        return [
            {
                "id": str(t.id),
                "first_name": t.first_name,
                "last_name": t.last_name,
                "email": t.email,
                "is_active": t.is_active,
                "profile_completed": t.profile_completed,
                "created_at": t.created_at,
            }
            for t in users
        ]

    # =========================
    # TEACHERS
    # =========================
    async def get_teachers(self, db, school_id: str):
        users = await self.repo.get_teachers(db, school_id)

        return [
            {
                "id": str(t.id),
                "first_name": t.first_name,
                "last_name": t.last_name,
                "email": t.email,
                "is_active": t.is_active,
                "profile_completed": t.profile_completed,
                "created_at": t.created_at,
            }
            for t in users
        ]

    async def update_student(
        self,
        db,
        school_id: str,
        student_id: str,
        payload: UpdateStudentRequest,
    ):
        student = await self.repo.get_student_by_id(
            db,
            school_id,
            student_id,
        )

        if not student:
            raise HTTPException(
                status_code=404,
                detail="Student not found",
            )

        student.first_name = payload.first_name
        student.last_name = payload.last_name
        student.email = payload.email

        await self.repo.commit(db)

        return {
            "message": "Student updated successfully",
        }

    async def update_teacher(
        self,
        db,
        school_id: str,
        teacher_id: str,
        payload: UpdateTeacherRequest,
    ):
        teacher = await self.repo.get_teacher_by_id(
            db,
            school_id,
            teacher_id,
        )

        if not teacher:
            raise HTTPException(
                status_code=404,
                detail="Teacher not found",
            )

        teacher.first_name = payload.first_name
        teacher.last_name = payload.last_name
        teacher.email = payload.email

        await self.repo.commit(db)

        return {
            "message": "Teacher updated successfully",
        }

    async def update_class(
        self,
        db,
        school_id: str,
        class_id: str,
        payload: UpdateClassRequest,
    ):
        school_class = await self.repo.get_class_by_id(
            db,
            school_id,
            class_id,
        )

        if not school_class:
            raise HTTPException(
                status_code=404,
                detail="Class not found",
            )

        school_class.name = payload.name
        school_class.level = payload.level

        await self.repo.commit(db)

        return {
            "message": "Class updated successfully",
        }

    async def update_subject(
        self,
        db,
        school_id: str,
        subject_id: str,
        payload: UpdateSubjectRequest,
    ):
        subject = await self.repo.get_subject_by_id(
            db,
            school_id,
            subject_id,
        )

        if not subject:
            raise HTTPException(
                status_code=404,
                detail="Subject not found",
            )

        subject.name = payload.name

        await self.repo.commit(db)

        return {
            "message": "Subject updated successfully",
        }

    async def delete_subject(
        self,
        db,
        school_id: str,
        subject_id: str,
    ):
        subject = await self.repo.get_subject_by_id(
            db,
            school_id,
            subject_id,
        )

        if not subject:
            raise HTTPException(
                status_code=404,
                detail="Subject not found",
            )

        await self.repo.delete_subject(
            db,
            subject,
        )

        await self.repo.commit(db)

        return {
            "message": "Subject deleted successfully",
        }

    async def export_results(
        self,
        db,
        school_id,
        session_id,
        term_id,
        class_id,
    ):
        batch = await self.result_repo.get_class_result_batch(
            db=db,
            school_id=school_id,
            class_id=class_id,
            session_id=session_id,
            term_id=term_id,
        )

        if not batch:
            raise HTTPException(
                status_code=404, detail="No results found for this class"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Class Results"

        # Professional title
        ws.merge_cells("A1:F1")
        ws["A1"] = "CLASS RESULT SHEET"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Summary headers
        headers = [
            "Student Name",
            "Position",
            "Average",
            "Total",
            "Passed",
            "Failed",
        ]

        header_row = 3
        ws.append([])
        ws.append(headers)

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid",
        )

        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        current_row = header_row + 1

        # Student summaries ordered by position
        summaries = sorted(
            batch.summaries,
            key=lambda s: (s.position if s.position is not None else 9999),
        )

        for summary in summaries:
            student_name = (
                f"{summary.student.first_name} {summary.student.last_name}"
                if summary.student
                else ""
            )

            # Summary row
            ws.cell(current_row, 1, student_name)
            ws.cell(current_row, 2, summary.position)
            ws.cell(current_row, 3, summary.average_score)
            ws.cell(current_row, 4, summary.total_score)
            ws.cell(current_row, 5, summary.passed_subjects)
            ws.cell(current_row, 6, summary.failed_subjects)

            for col in range(1, 7):
                cell = ws.cell(current_row, col)
                cell.border = border
                if col > 1:
                    cell.alignment = center

            # Emphasize student summary row
            for col in range(1, 7):
                ws.cell(current_row, col).font = Font(bold=True)

            summary_row = current_row
            current_row += 1

            # Subject detail header
            detail_header_row = current_row
            detail_headers = ["Subject", "CA", "Exam", "Total", "Grade", "Remark"]

            for idx, value in enumerate(detail_headers, start=1):
                cell = ws.cell(detail_header_row, idx + 1, value)  # start from column B
                cell.fill = PatternFill(
                    start_color="D9EAF7",
                    end_color="D9EAF7",
                    fill_type="solid",
                )
                cell.font = Font(bold=True, size=10)
                cell.alignment = center
                cell.border = border

            current_row += 1

            # Subject rows for this student
            student_records = [
                r for r in batch.records if r.student_id == summary.student_id
            ]

            start_group_row = current_row

            for record in student_records:
                ws.cell(current_row, 2, getattr(record.subject, "name", ""))
                ws.cell(current_row, 3, record.ca_score)
                ws.cell(current_row, 4, record.exam_score)
                ws.cell(current_row, 5, record.total_score)
                ws.cell(current_row, 6, record.grade)
                ws.cell(current_row, 7, record.remark)

                for col in range(2, 8):
                    cell = ws.cell(current_row, col)
                    cell.border = border
                    if col in [3, 4, 5, 6]:
                        cell.alignment = center

                current_row += 1

            end_group_row = current_row - 1

            # Group subject rows so Excel can expand/collapse them
            if end_group_row >= start_group_row:
                ws.row_dimensions.group(
                    start_group_row,
                    end_group_row,
                    hidden=False,
                    outline_level=1,
                )

            # Add spacing row
            current_row += 1

        # Freeze pane below headers
        ws.freeze_panes = "A4"

        # Auto-size columns
        for col_idx in range(1, 8):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in ws[column_letter]:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max_length + 4, 35)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "class_results.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    async def export_attendance(
        self,
        db,
        school_id: str,
        session_id: str,
        term_id: str,
    ):
        rows = await self.attendance_repo.get_attendance_analytics(
            db=db,
            school_id=school_id,
            session_id=session_id,
            term_id=term_id,
        )

        return rows


school_admin_service = SchoolAdminService()
