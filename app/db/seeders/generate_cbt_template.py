from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Output location inside Next.js public folder
output_dir = Path("public/templates")
output_dir.mkdir(parents=True, exist_ok=True)

output_file = output_dir / "cbt-question-template.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Questions"

headers = [
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_option",
    "marks",
]

# Header styling
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

ws.append(headers)

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Example rows
sample_rows = [
    [
        "What is 2 + 2?",
        "3",
        "4",
        "5",
        "6",
        "B",
        1,
    ],
    [
        "Capital of Nigeria?",
        "Lagos",
        "Abuja",
        "Kano",
        "Enugu",
        "B",
        2,
    ],
    [
        "Which planet is known as the Red Planet?",
        "Earth",
        "Venus",
        "Mars",
        "Jupiter",
        "C",
        1,
    ],
]

for row in sample_rows:
    ws.append(row)

# Set column widths
widths = {
    1: 45,  # question_text
    2: 20,
    3: 20,
    4: 20,
    5: 20,
    6: 18,
    7: 10,
}

for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Add instruction sheet
instructions = wb.create_sheet("Instructions")
instructions["A1"] = "CBT Question Upload Instructions"
instructions["A1"].font = Font(bold=True, size=14)

instruction_lines = [
    "1. Do not change the column headers in the Questions sheet.",
    "2. Enter one question per row.",
    "3. correct_option must be A, B, C, or D.",
    "4. marks must be a positive whole number.",
    "5. Save the file as .xlsx and upload it through the Batch Upload dialog.",
]

for idx, line in enumerate(instruction_lines, start=3):
    instructions[f"A{idx}"] = line

instructions.column_dimensions["A"].width = 100

wb.save(output_file)

print(f"Template created: {output_file.resolve()}")
